import openpyxl
from datetime import date
from gooey import Gooey, GooeyParser

@Gooey(program_name="Blueberry Barcode Configurator")
def parse_args():
    parser = GooeyParser(description="Make sure that the input file includes columns entitled Sel, Year, and Stage: .xlsx only")
    #subparsers = parser.add_subparsers(required = True)
    #place = subparsers.add_parser("Select Site", help = "Where will these samples be harvested?")

    parser.add_argument("File",widget="FileChooser",help="File to append")
    parser.add_argument('Select Site', choices=['Waldo', 'Citra','Windsor'], help='Harvest Location', widget='Dropdown')
    return parser.parse_args()


args = parse_args()
args = str(args)
print(args)
#print(lword)
start = args.find("=") + 2
end = args.find(",") -1
path = args[start:end]
start = args.find(":") + 3
end = args.find("}") - 1
site = args[start:end]
print(path)
print(site)
#path = "samples.xlsx"



wb = openpyxl.load_workbook(path)
sheet = wb.active

#site = "%s" % ("".join(site.upper().split(" ")))
obj = sheet.cell(row = 1, column = 20)
obj.value = "Time"
row_count = sheet.max_row
today = date.today()
time = today.strftime("%b/%d/%y")

for i in range(2, row_count+1):
    obj = sheet.cell(row=i, column=20)
    obj.value = time

header1 = sheet.cell(row = 1, column = 21)
header1.value="Barcode String"

for columns in sheet.iter_cols(1):
    for cell in columns:
        if cell.value == "Sel":
            genoCol = cell.column

        if cell.value == "Year":
            BlockYearCol = cell.column

        if cell.value == "Stage":
            stageCol = cell.column


#print(genoCol)
#print(BlockYearCol)
#print(stageCol)

for i in range(2, row_count+1):
    geno = sheet.cell(row=i,column = genoCol)
    blockYear = sheet.cell(row=i,column = BlockYearCol)
    stage = sheet.cell(row=i,column = stageCol)
    barString = "GENO:" + str(geno.value) + "SITE:" + site + "BLOCKYR:" + str(blockYear.value) +"STG:" + str(stage.value) + "PROJ:" + "none" + "NOTE:" + "none" + "TIME:" + str(time)
    obj = sheet.cell(row=i, column=21)
    obj.value = barString


wb.save(path)
